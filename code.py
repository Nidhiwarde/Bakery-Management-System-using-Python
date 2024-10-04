import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import smtplib
from email.mime.text import MIMEText
from twilio.rest import Client
import datetime

# Twilio credentials (replace with your own)
TWILIO_ACCOUNT_SID = 'AC21de5885206aa6658e661c708fc05089'
TWILIO_AUTH_TOKEN = 'e55c37a36a2738729221a4e5b6b48f6c'
TWILIO_PHONE_NUMBER = '+12513135757'
OWNER_PHONE_NUMBER = '+918591720424'

# Email credentials (replace with your own)
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587
EMAIL_ADDRESS = 'wardenidhi999@gmail.com'
EMAIL_PASSWORD = 'okin cqfc kosf lqwp'
OWNER_EMAIL = 'nidhiwarde26@gmail.com'

# Create or load the Excel workbook
excel_file = 'bakery_orders.xlsx'
try:
    wb = load_workbook(excel_file)
    sheet = wb.active
except FileNotFoundError:
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Order ID", "Customer Name", "Order Time", "Billing Amount", "Payment Done", "Mobile Number", "Email ID"])

class BakeryManagementSystem(tk.Tk):
    def _init_(self):
        super()._init_()
        self.title("Bakery Management System")
        self.geometry("400x400")

        self.create_widgets()

    def create_widgets(self):
        tk.Label(self, text="Customer Name:").pack(pady=5)
        self.customer_name_entry = tk.Entry(self)
        self.customer_name_entry.pack(pady=5)

        tk.Label(self, text="Billing Amount:").pack(pady=5)
        self.billing_amount_entry = tk.Entry(self)
        self.billing_amount_entry.pack(pady=5)

        tk.Label(self, text="Payment Done (yes/no):").pack(pady=5)
        self.payment_done_entry = tk.Entry(self)
        self.payment_done_entry.pack(pady=5)

        tk.Label(self, text="Mobile Number:").pack(pady=5)
        self.mobile_number_entry = tk.Entry(self)
        self.mobile_number_entry.pack(pady=5)

        tk.Label(self, text="Email ID:").pack(pady=5)
        self.email_id_entry = tk.Entry(self)
        self.email_id_entry.pack(pady=5)

        tk.Button(self, text="Place Order", command=self.place_order).pack(pady=20)

    def place_order(self):
        customer_name = self.customer_name_entry.get()
        billing_amount = self.billing_amount_entry.get()
        payment_done = self.payment_done_entry.get()
        mobile_number = self.mobile_number_entry.get()
        email_id = self.email_id_entry.get()

        if not customer_name or not billing_amount or not payment_done or not mobile_number or not email_id:
            messagebox.showerror("Error", "All fields are required")
            return

        order_id = len(sheet['A'])  # Order ID is the next row number
        order_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        sheet.append([order_id, customer_name, order_time, billing_amount, payment_done, mobile_number, email_id])
        wb.save(excel_file)

        order_details = f"""
        Order ID: {order_id}
        Customer Name: {customer_name}
        Order Time: {order_time}
        Billing Amount: {billing_amount}
        Payment Done: {payment_done}
        Mobile Number: {mobile_number}
        Email ID: {email_id}
        """

        self.send_email(email_id, order_details)
        self.send_sms(mobile_number, order_details)
        self.send_sms(OWNER_PHONE_NUMBER, order_details)

        messagebox.showinfo("Success", "Order placed successfully!")

    def send_email(self, recipient_email, order_details):
        msg = MIMEText(order_details)
        msg['Subject'] = 'Order Confirmation'
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = recipient_email

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, recipient_email, msg.as_string())
            server.sendmail(EMAIL_ADDRESS, OWNER_EMAIL, msg.as_string())

    def send_sms(self, recipient_number, order_details):
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        client.messages.create(
            body=order_details,
            from_=TWILIO_PHONE_NUMBER,
            to=recipient_number
        )

if _name_ == "_main_":
    app = BakeryManagementSystem()
    app.mainloop()
